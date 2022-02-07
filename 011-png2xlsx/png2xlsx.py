import os
from PIL import Image
import openpyxl
from openpyxl.styles import PatternFill

pngPath = ''
while (not os.path.isfile(pngPath)):
    pngPath = input('Input PNG   = ')

png = Image.open(pngPath)
print('PNG Size    = ' + str(png.width) + 'x' + str(png.height))

scale = input('Output % (default 1) = ')
try:
    scale = int(scale)
except ValueError:
    scale = 1

output = input('Output XLSX          = ')

NTFSREPLACEMENT = ['\\', '/', ':', '*', '?', '"', '<', '>', '|']
if output == '':
    output = pngPath
    for char in NTFSREPLACEMENT:
        if char in output:
            output = output.replace(char, '')
    output = 'png2xslx-'+output+'.xlsx'


# Resize image
print('Scaling PNG')
scale = max(1, min(scale, 100))/100
png = png.convert("RGBA")
png = png.resize((int(png.width * scale), int(png.height * scale)))
pixel = png.load()

xls = openpyxl.Workbook()
ws = xls.active
ws.title = output

# Color everything
print('Coloring XLSX')
for y in range(png.height):
    for x in range(png.width):
        ThisCell = ws.cell(row=y+1, column=x+1)
        r, g, b, a = pixel[x,y]
        ThisFill = f"{a:02x}{r:02x}{g:02x}{b:02x}"

        ThisCell.fill = PatternFill(fgColor=ThisFill, fill_type="solid")

# Resize cells to square pixels
print('Resizing XLSX')
for col in range(png.width):
    ws.column_dimensions[openpyxl.utils.cell.get_column_letter(col+1)].width = 1

for row in range(png.height):
    ws.row_dimensions[row+1].height = 5.25

# End
xls.save(output)
print('Saved as ' + output)
exit(0)